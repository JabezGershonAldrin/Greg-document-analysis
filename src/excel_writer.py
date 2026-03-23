import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def write_excel(data, output_file):

    df = pd.DataFrame(data)

    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(output_file)